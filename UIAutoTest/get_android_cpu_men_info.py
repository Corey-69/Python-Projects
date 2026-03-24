import uiautomator2 as u2
import pandas as pd
import time
import datetime
import re
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series


# ================= 数据处理与计算函数 =================

def calculate_cpu_percent(prev_idle, prev_total, curr_idle, curr_total):
    """计算 CPU 占用率"""
    total_diff = curr_total - prev_total
    idle_diff = curr_idle - prev_idle
    if total_diff == 0:
        return 0.0
    return (total_diff - idle_diff) / total_diff * 100


def parse_proc_stat(stat_output):
    """解析 /proc/stat 获取CPU占用率"""
    lines = stat_output.strip().split('\n')
    first_line = lines[0]
    if not first_line.startswith('cpu '):
        raise ValueError("Invalid /proc/stat format")

    parts = first_line.split()
    values = [int(x) for x in parts[1:]]

    # user(0), nice(1), system(2), idle(3), iowait(4)...
    idle = values[3]
    current_total = sum(values)
    return idle, current_total


def get_memory_usage(d):
    """解析 /proc/meminfo 获取内存占用率"""
    output = d.shell("cat /proc/meminfo").output
    mem_info = {}

    for line in output.splitlines():
        if ':' in line:
            key, value = line.split(':')
            key = key.strip()
            num_str = re.sub(r'[^0-9]', '', value.split()[0])
            if num_str:
                mem_info[key] = int(num_str)

    total = mem_info.get('MemTotal', 0)
    free = mem_info.get('MemFree', 0)
    buffers = mem_info.get('Buffers', 0)
    cached = mem_info.get('Cached', 0)

    available = free + buffers + cached
    used = total - available

    if total == 0:
        return 0, 0.0
    percent = (used / total) * 100
    return total, percent


# ================= 图表生成函数 =================

def create_chart(df, filename):
    """
    在 Excel 文件中添加第二个 Sheet，仅放置图表
    """
    print("正在生成图表...")
    try:
        # 加载已有的 workbook
        wb = load_workbook(filename)

        # 如果已存在"图表分析" sheet，则删除它以便重新生成
        if "图表分析" in wb.sheetnames:
            del wb["图表分析"]

        ws_chart = wb.create_sheet("图表分析")

        # 为了图表能引用数据，我们需要将源数据（Sheet1）的引用指向第一个Sheet
        # 而不是在当前Sheet重复写入数据。
        # 获取第一个Sheet的名字（通常是"性能数据"）
        source_sheet_name = wb.sheetnames[0]
        ws_source = wb[source_sheet_name]

        max_row = len(df) + 1  # 包含标题行

        # 创建折线图
        chart = LineChart()
        chart.title = "性能监控趋势图 (CPU & 内存)"
        chart.style = 13
        chart.y_axis.title = "占用率 (%)"
        chart.x_axis.title = "时间戳"

        # 定义数据范围 (引用第一个 Sheet 的数据)
        # CPU 数据：Sheet1 的 B 列 (第2列)
        # Reference(sheet, min_col, min_row, max_col, max_row)
        cpu_data = Reference(ws_source, min_col=2, min_row=1, max_row=max_row)
        cpu_series = Series(cpu_data, title_from_data=True)
        chart.series.append(cpu_series)

        # 内存 数据：Sheet1 的 C 列 (第3列)
        mem_data = Reference(ws_source, min_col=3, min_row=1, max_row=max_row)
        mem_series = Series(mem_data, title_from_data=True)
        chart.series.append(mem_series)

        # 设置 X 轴标签 (引用第一个 Sheet 的 A 列)
        cats = Reference(ws_source, min_col=1, min_row=2, max_row=max_row)
        chart.set_categories(cats)

        # 设置图表尺寸 (放大尺寸)
        # width 和 height 单位大约是像素/点数，默认通常较小，这里设置为更大
        chart.width = 25  # 默认约 7-8，设置为 25 更宽
        chart.height = 15  # 默认约 5-6，设置为 15 更高

        # 将图表添加到当前 Sheet 的 A1 位置，居中展示
        ws_chart.add_chart(chart, "A1")

        # 保存文件
        wb.save(filename)
        print(f"✅ 图表已生成并保存至文件: {filename} 的 [图表分析] 页签")

    except Exception as e:
        print(f"❌ 生成图表失败: {e}")
        print("提示: 请确保已安装 openpyxl 库 (pip install openpyxl)")


# ================= 主监控逻辑 =================

def monitor_performance(serial=None, duration=30, interval=3):
    print(f"正在连接设备...")
    try:
        if serial:
            d = u2.connect(serial)
        else:
            d = u2.connect()
        print(f"连接成功: {d.device_info.get('serial', 'Unknown')}")
    except Exception as e:
        print(f"连接失败: {e}")
        return

    # 获取设备序列号用于命名文件
    try:
        device_name = d.adb_device.serial.replace(":", "_")
    except:
        device_name = "android_device"

    data_list = []
    start_time = time.time()
    output_file_name = f"{device_name}-性能数据.xlsx"

    # --- 第一次采样 (预热) ---
    print("正在进行首次 CPU 采样预热...")
    try:
        stat_out = d.shell("cat /proc/stat").output
        prev_idle, prev_total = parse_proc_stat(stat_out)
    except Exception as e:
        print(f"读取 CPU 信息失败: {e}")
        return

    print(f"开始监控，持续 {duration} 秒，间隔 {interval} 秒...")

    while (time.time() - start_time) < duration:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 1. 获取 CPU
        try:
            stat_out = d.shell("cat /proc/stat").output
            curr_idle, curr_total = parse_proc_stat(stat_out)
            cpu_percent = calculate_cpu_percent(prev_idle, prev_total, curr_idle, curr_total)
            prev_idle, prev_total = curr_idle, curr_total
        except Exception as e:
            print(f"CPU 采集错误: {e}")
            cpu_percent = 0.0

        # 2. 获取内存
        try:
            _, mem_percent = get_memory_usage(d)
        except Exception as e:
            print(f"内存采集错误: {e}")
            mem_percent = 0.0

        # 记录
        row = {
            "时间戳": timestamp,
            "CPU占用率": round(cpu_percent, 2),
            "内存占用率": round(mem_percent, 2)
        }
        data_list.append(row)
        print(f"[{timestamp}] CPU: {row['CPU占用率']:>6.2f}% | Mem: {row['内存占用率']:>6.2f}%")

        time.sleep(interval)

    # --- 数据导出与绘图 ---
    if data_list:
        df = pd.DataFrame(data_list)

        # 1. 先保存原始数据到 Sheet1
        try:
            # index=False 不保存行号
            df.to_excel(output_file_name, sheet_name="性能数据", index=False)
            print(f"\n✅ 原始数据已保存至: {output_file_name}")

            # 2. 调用绘图函数，在同一文件生成 Sheet2 (仅图表)
            create_chart(df, output_file_name)

        except Exception as e:
            print(f"保存文件失败: {e}")
            # 降级保存 CSV
            csv_file = output_file_name.replace('.xlsx', '.csv')
            df.to_csv(csv_file, index=False)
            print(f"已降级保存为 CSV: {csv_file}")
    else:
        print("未采集到任何数据。")


if __name__ == "__main__":
    # 配置区域
    DEVICE_SERIAL = None  # 自动连接
    DURATION = 60 * 60 * 24 * 7  # 监控 60 秒
    INTERVAL = 3  # 每 3 秒一次

    monitor_performance(
        serial=DEVICE_SERIAL,
        duration=DURATION,
        interval=INTERVAL
    )